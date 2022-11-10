from fafscripts.modules import notion_new as n, utils as u, dbfuncs
import openpyxl as xl
from openpyxl.styles import Font, PatternFill
from random import randrange
from fafscripts.models import EventCategory
import logging


def main(folder_location):
    logger = logging.getLogger(__name__)
    logger.info(f"Starting {__name__}")
    db_identifier_to_date = {
        0: 'Thursday, 20.10',
        1: 'Friday, 21.10',
        2: 'Saturday, 22.10',
        3: 'Sunday, 23.10',
    }

    # basic setup
    event_dbs = list()
    bold = Font(name='Calibri', sz=12, bold=True)
    fill_red = PatternFill(start_color="FFC7CE", fill_type="solid")
    fill_yellow = PatternFill(start_color="00FFFF00", fill_type="solid")
    fill_green = PatternFill(start_color="0099CC00", fill_type="solid")

    # retrieve venue db
    notion_venues = n.get_db('venues')

    # retrieve filtered and sorted databases for festival days, storing them in event_dbs list
    # OBS has to be updated
    logger.info("Retrieving databases based on 'Time'")
    data_dict = {}
    n.add_filter_to_request_dict(
        data_dict, 'Time', 'date', 'equals', "2022-10-20")
    n.add_sorts_to_request_dict(data_dict, 'Time', 'ascending')
    events_thursday = n.get_db('events', data_dict=data_dict)
    event_dbs.append(events_thursday)

    data_dict = {}
    n.add_filter_to_request_dict(
        data_dict, 'Time', 'date', 'equals', "2022-10-21")
    n.add_sorts_to_request_dict(data_dict, 'Time', 'ascending')
    events_friday = n.get_db('events', data_dict=data_dict)
    event_dbs.append(events_friday)

    data_dict = {}
    n.add_filter_to_request_dict(
        data_dict, 'Time', 'date', 'equals', "2022-10-22")
    n.add_sorts_to_request_dict(data_dict, 'Time', 'ascending')
    events_saturday = n.get_db('events', data_dict=data_dict)
    event_dbs.append(events_saturday)

    data_dict = {}
    n.add_filter_to_request_dict(
        data_dict, 'Time', 'date', 'equals', "2022-10-23")
    n.add_sorts_to_request_dict(data_dict, 'Time', 'ascending')
    events_sunday = n.get_db('events', data_dict=data_dict)
    event_dbs.append(events_sunday)

    # create spreadsheet
    logger.info('Creating schedule spreadsheet...')
    wb = xl.Workbook()
    sheet = wb.active

    line_count = 0
    db_identifier = 0

    # looping through event_dbs list
    for event_db in event_dbs:
        line_count += 1
        sheet.cell(line_count, 1).value = db_identifier_to_date.get(
            db_identifier)
        sheet.cell(line_count, 1).font = bold
        db_identifier += 1
        line_count += 1
        sheet.cell(line_count, 1).value = 'Start'
        sheet.cell(line_count, 2).value = 'End'
        sheet.cell(line_count, 3).value = 'Event name'
        sheet.cell(line_count, 4).value = 'Info'
        sheet.cell(line_count, 5).value = 'Screen/Venue'
        # looping through events
        for event_json in event_db['results']:
            e = n.Page(json_obj=event_json)
            type_id = e.get_list('category')[0]
            event_type = dbfuncs.get_name_from_notion_id(
                type_id, EventCategory)
            hide = e.get_text('hide-from-catalogue')
            if hide:
                continue
            if event_type == "VIP event":  # Filter out 'VIP events"
                continue

            line_count += 1
            # fill in text
            try:
                sheet.cell(line_count, 1).value = e.get_date('time')[0][11:16]
            except IndexError:
                raise IndexError(
                    "Events need to have a proper start time for schedule script to work")

            try:
                sheet.cell(line_count, 2).value = e.get_date('time')[1][11:16]
            except (TypeError, IndexError):
                logger.warning(f'OBS: Encountered missing end time for event.')
            sheet.cell(line_count, 3).value = e.get_text('name')
            other_info = e.get_list('other-info')
            if len(other_info) > 0:
                other_info = u.list_to_comma_separated(
                    other_info).replace(',', " /")
            else:
                other_info = ""
            age_limit = e.get_text('age-limit')
            if age_limit:
                # if age_limit == 0:
                #     age_limit = "A"
                if other_info:
                    # dealing with annoying space
                    other_info = f"{other_info} "
                other_info = f"{other_info}(Age limit {age_limit})"
            sheet.cell(line_count, 4).value = other_info
            venue_notion_id = e.get_list('venue')
            try:
                venue_notion_page = n.get_page_from_db(
                    venue_notion_id[0], source=notion_venues)
            except IndexError:
                raise IndexError(
                    "Events need to have a venue assigned for catalogue schedule script to work")
            v = n.Page(json_obj=venue_notion_page)
            sheet.cell(line_count, 5).value = v.get_text('name')
            # apply formatting  OBS event types/names

            if event_type == 'Free Event' or event_type == 'Ticketed Event':
                sheet.cell(line_count, 3).fill = fill_red
                sheet.cell(line_count, 4).fill = fill_red
            elif 'Screening' in event_type:
                sheet.cell(line_count, 3).fill = fill_green
                sheet.cell(line_count, 4).fill = fill_green
            elif event_type == 'Industry':
                sheet.cell(line_count, 3).fill = fill_yellow
                sheet.cell(line_count, 4).fill = fill_yellow

    # add type information at bottom of spreadsheet
    line_count += 2
    sheet.cell(line_count, 3).value = 'Screenings'
    sheet.cell(line_count, 3).fill = fill_green
    line_count += 1
    sheet.cell(line_count, 3).value = 'Industry'
    sheet.cell(line_count, 3).fill = fill_yellow
    line_count += 1
    sheet.cell(line_count, 3).value = 'Events'
    sheet.cell(line_count, 3).fill = fill_red

    # save worksheet
    rand = randrange(1000)
    wb.save(f"temp{rand}.xlsx")
    u.dropbox_upload_local_file(
        f"temp{rand}.xlsx", f"{folder_location}/schedule.xlsx")
    logger.info('Schedule saved.')
