"""Create a spreadsheet (almost) ready to be sent in, download invoices/receipts with appropriate filenames
(and combine them?)
"""

from fafscripts.modules import notion_new as n, utils as u
import openpyxl as xl
from urllib import request
from forex_python.converter import CurrencyRates  # pip install forex-python
from datetime import date
from datetime import datetime
from random import randrange
import logging


def main(name: str, address, account_number, id_number, export_files=False):
    # BASIC SETUP
    logger = logging.getLogger(__name__)
    folder_location = f"{u.get_secret('DROPBOX_FOLDER')}/receipts_{name}"
    festival_address = u.get_secret("FESTIVAL_ADDRESS")

# RETRIEVE NOTION DB
    request_dict = dict()
# filter for 'Reimburse' (person to be reimbursed)
# sort by receipt date
    n.add_filter_to_request_dict(
        request_dict, 'Reimburse', 'select', 'equals', name)
    n.add_sorts_to_request_dict(request_dict, 'Date')
    notion_receipts = n.get_db('receipts', data_dict=request_dict)

# RETRIEVE CURRENCY EXCHANGE RATES AND CURRENT DATE
    c = CurrencyRates()
    euros_to_nok = round(c.get_rate('EUR', 'NOK'), 2)
    dollars_to_nok = round(c.get_rate('USD', 'NOK'), 2)
    # euros_to_nok = 10.36
    # dollars_to_nok = 10.68
    today = date.today()

# INITIALIZE SPREADSHEET
    wb = xl.Workbook()
    sheet = wb.active
    line = 1
# LINE W/DATE (c6)
    sheet.cell(line, 6).value = today
    line += 1
# BLANK LINE
    line += 1
# THREE LINES W/ P. ADDRESS (c6)

    address_split = address.split('\n')
    sheet.cell(line, 6).value = address_split[0]
    line += 1
    sheet.cell(line, 6).value = address_split[1]
    line += 1
    sheet.cell(line, 6).value = address_split[2]
    line += 1
# BLANK LINE
    line += 1
# 4 LINES ADDRESS(c1)
    festival_address_split = festival_address.split('\n')
    sheet.cell(line, 1).value = festival_address_split[0]
    line += 1
    sheet.cell(line, 1).value = festival_address_split[1]
    line += 1
    sheet.cell(line, 1).value = festival_address_split[2]
    line += 1
    sheet.cell(line, 1).value = festival_address_split[3]
    line += 1
# BLANK LINE
    line += 1
# Title (c1) add info on currency exchange rate (w/ date) (c4)
    sheet.cell(
        line, 1).value = f"Expenses in conjunction with {u.get_secret('FESTIVAL_ACRONYM')}"
    sheet.cell(
        line, 4).value = f'1€={euros_to_nok} NOK, 1$={dollars_to_nok} NOK (as of {today})'
    line += 1
# TWO BLANK LINES
    line += 2
# LINE W/ HEADINGS #-1 Date-2 Name-3 In NOK-4 In original currency-5
    sheet.cell(line, 1).value = '#'
    sheet.cell(line, 2).value = 'Date'
    sheet.cell(line, 3).value = 'What?'
    sheet.cell(line, 4).value = 'In NOK'
    sheet.cell(line, 5).value = 'In currency'
    line += 1

    total = 0
    # receipt_count = 1
# LOOP THROUGH RECEIPTS
    for index, receipt_json in enumerate(notion_receipts['results']):
        r = n.Page(json_obj=receipt_json)
        # skip 'paid'
        if r.get_text('paid'):
            continue
        # load in required information
        r_date = r.get_date('date')
        date_string = u.get_date_string(r_date[0], output='day')
        # test
        test_date = datetime.fromisoformat(r_date[0])
        date_stamp = test_date.strftime("%Y%m%d")
        # / test
        # OBS test numbers behaviour
        currency = r.get_text('currency')
        amount = r.get_text('amount')
        title = r.get_text('name')
        file_url = r.get_list('file')
        if not file_url:
            logger.warning(f"No file found for {title}.")

        # add line in spreadsheet
        # ... Seq (#), Name, Currency, amount in Nok
        sheet.cell(line, 1).value = str(index + 1).zfill(2)
        sheet.cell(line, 2).value = date_string  # Date
        sheet.cell(line, 3).value = title  # What
        if currency == '€':
            amount_in_nok = round(amount * euros_to_nok, 2)
        elif currency == '$':
            amount_in_nok = round(amount * dollars_to_nok, 2)
        else:
            amount_in_nok = amount
        sheet.cell(line, 4).value = f'kr {amount_in_nok}'  # in NOK
        if currency != 'kr':
            # In original currency
            sheet.cell(line, 5).value = f'{amount} {currency}'
        # add amount to total
        total += amount_in_nok
        # SAVE FILE W/ CORRECT FILE NAME Seq.zfill(2)_YYYYMMDD_NameWithoutSpaces
        file_extension = 'pdf' if '.pdf' in file_url[0] else 'jpg'
        # add pdf convert?
        if export_files:
            filename = f"{folder_location}/{str(index + 1).zfill(2)}_{date_stamp}_{title.title().replace(' ','')}.{file_extension}"
            rand = randrange(1000)
            request.urlretrieve(file_url[0], f"temp{rand}.{file_extension}")
            u.dropbox_upload_local_file(
                f"temp{rand}.{file_extension}", filename)

        # increment line count
        line += 1

# BLANK LINE
    line += 1
# ADD LINE FOR SUM (c3/4)
    sheet.cell(line, 3).value = 'Σ'
    sheet.cell(line, 4).value = f"kr {round(total,2)}"
    line += 1
# BLANK LINE
    line += 1
# LINE W/ NAME, F-NUMBER
    sheet.cell(line, 1).value = 'Payable to:'
    sheet.cell(line, 3).value = address_split[0]
    sheet.cell(line, 4).value = id_number
    line += 1
# LINE W/ ACCOUNT NUMBER
    sheet.cell(line, 3).value = account_number

# SAVE SPREADSHEET
    rand = randrange(1000)
    wb.save(f'temp{rand}.xlsx')
    u.dropbox_upload_local_file(
        f"temp{rand}.xlsx", f"{folder_location}/reimbursements_{name}.xlsx")
