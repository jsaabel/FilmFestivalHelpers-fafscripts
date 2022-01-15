import modules.utils
from modules import notion as n, utils as u
import openpyxl as xl
from openpyxl.styles import Font, PatternFill
from time import sleep


def main(folder_location:str = "../exports"):
    db_identifier_to_date = {
        0: 'Thursday, 21.10',
        1: 'Friday, 22.10',
        2: 'Saturday, 23.10',
        3: 'Sunday, 24.10',
    }

    # basic setup
    event_dbs = list()
    bold = Font(name='Calibri', sz=12, bold=True)
    fill_red = PatternFill(start_color="FFC7CE", fill_type="solid")
    fill_yellow = PatternFill(start_color="00FFFF00", fill_type="solid")
    fill_green = PatternFill(start_color="0099CC00", fill_type="solid")

    # retrieve venue db
    notion_venues = n.get_db('venues')

    # retrieve filtered and sorted databases for festival days, storing them in event_dbs list
    print('Retrieving databases...')
    data_dict = {}
    n.add_filter_to_request_dict(data_dict, 'Time', 'date', 'equals', "2021-10-21T00:00:00Z")
    n.add_sorts_to_request_dict(data_dict, 'Time', 'ascending')
    events_thursday = n.get_db('events', data_dict=data_dict)
    event_dbs.append(events_thursday)

    data_dict = {}
    n.add_filter_to_request_dict(data_dict, 'Time', 'date', 'equals', "2021-10-22T00:00:00Z")
    n.add_sorts_to_request_dict(data_dict, 'Time', 'ascending')
    events_friday = n.get_db('events', data_dict=data_dict)
    event_dbs.append(events_friday)

    data_dict = {}
    n.add_filter_to_request_dict(data_dict, 'Time', 'date', 'equals', "2021-10-23T00:00:00Z")
    n.add_sorts_to_request_dict(data_dict, 'Time', 'ascending')
    events_saturday = n.get_db('events', data_dict=data_dict)
    event_dbs.append(events_saturday)

    data_dict = {}
    n.add_filter_to_request_dict(data_dict, 'Time', 'date', 'equals', "2021-10-24T00:00:00Z")
    n.add_sorts_to_request_dict(data_dict, 'Time', 'ascending')
    events_sunday = n.get_db('events', data_dict=data_dict)
    event_dbs.append(events_sunday)

    # create spreadsheet
    print('Creating schedule spreadsheet...')
    wb = xl.Workbook()
    sheet = wb.active

    line_count = 0
    db_identifier = 0

    # looping through event_dbs list
    for event_db in event_dbs:
        line_count += 1
        sheet.cell(line_count, 1).value = db_identifier_to_date[db_identifier]
        sheet.cell(line_count, 1).font = bold
        db_identifier += 1
        line_count += 1
        sheet.cell(line_count, 1).value = 'Start'
        sheet.cell(line_count, 2).value = 'End'
        sheet.cell(line_count, 3).value = 'Event name'
        sheet.cell(line_count, 4).value = 'Info'
        sheet.cell(line_count, 5).value = 'Screen/Venue'
        # looping through events
        item_count_notion = n.get_item_count(event_db)
        for index in range(0, item_count_notion):
            event_type = n.get_property(index, 'Type', 'select', source=event_db)
            if event_type == "VIP event":  # Filter out 'VIP events"
                pass
            else:
                line_count += 1
                # fill in text
                sheet.cell(line_count, 1).value = n.get_property(index, 'Time', 'date', source=event_db)[0][11:16]
                try:
                    sheet.cell(line_count, 2).value = n.get_property(index, 'Time', 'date', source=event_db)[1][11:16]
                except TypeError:
                    print(f'OBS: Encountered missing end time for event.')
                sheet.cell(line_count, 3).value = n.get_property(index, 'Name', 'title', source=event_db)
                other_info = n.get_property(index, 'Other Info', 'multi_select', source=event_db)
                other_info = u.list_to_comma_separated(other_info).replace(','," /")
                age_limit = n.get_property(index, 'Age Limit', 'number', source=event_db)
                if age_limit:
                    # if age_limit == 0:
                    #     age_limit = "A"
                    if other_info:
                        other_info = f"{other_info} " # dealing with annoying space
                    other_info = f"{other_info}(Age limit {age_limit})"
                sheet.cell(line_count, 4).value = other_info
                venue_notion_id = n.get_property(index, 'Venue', 'relation', source=event_db)
                venue_notion_page = n.get_page_from_db(venue_notion_id[0], source=notion_venues)
                venue = n.get_property_from_page('Name', 'title', source=venue_notion_page)
                sheet.cell(line_count, 5).value = venue
                # apply formatting  OBS event types/names

                if event_type == 'Free Event' or event_type == 'Ticketed Event':
                    sheet.cell(line_count, 3).fill = fill_red
                    sheet.cell(line_count, 4).fill = fill_red
                elif 'Screening' in event_type:
                    sheet.cell(line_count, 3).fill = fill_green
                    sheet.cell(line_count, 4).fill = fill_green
                else:
                    sheet.cell(line_count, 3).fill = fill_yellow
                    sheet.cell(line_count, 4).fill = fill_yellow

    # add type information at bottom of spreadsheet
    line_count += 2
    sheet.cell(line_count, 3).value = 'Screenings'
    sheet.cell(line_count, 3).fill = fill_green
    line_count += 1
    sheet.cell(line_count, 3).value = 'Industry'
    sheet.cell(line_count, 3).fill = fill_yellow
    line_count += 1
    sheet.cell(line_count, 3).value = 'Events'
    sheet.cell(line_count, 3).fill = fill_red

    # save worksheet
    file_name = f"{folder_location}/schedule.xlsx"
    wb.save(file_name)
    print('Schedule saved.')

