"""Create a spreadsheet (almost) ready to be sent in, download invoices/receipts with appropriate filenames
(and combine them?)
"""

from modules import notion as n, utils as u
import openpyxl as xl
from urllib import request
from forex_python.converter import CurrencyRates  # pip install forex-python
from datetime import date
from datetime import datetime
import os

# BASIC SETUP
folder_location = 'receipts'
try:
    os.makedirs(folder_location)
except FileExistsError:
    pass

# RETRIEVE NOTION DB
request_dict = dict()
# filter for 'Reimburse' (person to be reimbursed)
# sort by receipt date
n.add_filter_to_request_dict(request_dict, 'Reimburse', 'select', 'equals', 'Jonas')
n.add_sorts_to_request_dict(request_dict, 'Date')
notion_receipts = n.get_db('receipts', data_dict=request_dict)
item_count_notion = n.get_item_count(notion_receipts)

# RETRIEVE CURRENCY EXCHANGE RATES AND CURRENT DATE
c = CurrencyRates()
euros_to_nok = round(c.get_rate('EUR', 'NOK'), 2)
dollars_to_nok = round(c.get_rate('USD', 'NOK'), 2)
today = date.today()

# INITIALIZE SPREADSHEET
wb = xl.Workbook()
sheet = wb.active
line = 1
# LINE W/DATE (c6)
sheet.cell(line, 6).value = today
line += 1
# BLANK LINE
line += 1
# THREE LINES W/ P. ADDRESS (c6)
sheet.cell(line, 6).value = '...'
line += 1
sheet.cell(line, 6).value = '...'
line += 1
sheet.cell(line, 6).value = '...'
line += 1
# BLANK LINE
line += 1
# 4 LINES ADDRESS(c1)
sheet.cell(line, 1).value = '...'
line += 1
sheet.cell(line, 1).value = '...'
line += 1
sheet.cell(line, 1).value = '...'
line += 1
sheet.cell(line, 1).value = '...'
line += 1
# BLANK LINE
line += 1
# Title (c1) add info on currency exchange rate (w/ date) (c4)
sheet.cell(line, 1).value = 'Expenses in conjunction with ...'
sheet.cell(line, 4).value = f'1€={euros_to_nok} NOK, 1$={dollars_to_nok} NOK (as of {today})'
line += 1
# TWO BLANK LINES
line += 2
# LINE W/ HEADINGS #-1 Date-2 Name-3 In NOK-4 In original currency-5
sheet.cell(line, 1).value = '#'
sheet.cell(line, 2).value = 'Date'
sheet.cell(line, 3).value = 'What?'
sheet.cell(line, 4).value = 'In NOK'
sheet.cell(line, 5).value = 'In currency'
line += 1

total = 0
receipt_count = 1
# LOOP THROUGH RECEIPTS
for index in range(item_count_notion):
    # load in required information
    r_date = n.get_property(index, 'Date', 'date', source=notion_receipts)
    date_string = u.get_date_string(r_date[0], output='day')
    # test
    test_date = datetime.fromisoformat(r_date[0])
    date_stamp = test_date.strftime("%Y%m%d")
    # / test
    currency = n.get_property(index, 'currency', 'select', source=notion_receipts)
    amount = n.get_property(index, 'amount', 'number', source=notion_receipts)
    title = n.get_property(index, 'Name', 'title', source=notion_receipts)
    file_url = n.get_property(index, 'File', 'files', source=notion_receipts)[0]

    # add line in spreadsheet
    # ... Seq (#), Name, Currency, amount in Nok
    sheet.cell(line, 1).value = str(index + 1).zfill(2)
    sheet.cell(line, 2).value = date_string  # Date
    sheet.cell(line, 3).value = title  # What
    if currency == '€':
        amount_in_nok = round(amount * euros_to_nok, 2)
    elif currency == '$':
        amount_in_nok = round(amount * dollars_to_nok, 2)
    else:
        amount_in_nok = amount
    sheet.cell(line, 4).value = f'kr {amount_in_nok}'  # in NOK
    if currency != 'kr':
        sheet.cell(line, 5).value = f'{amount} {currency}'  # In original currency
    # add amount to total
    total += amount_in_nok
    # SAVE FILE W/ CORRECT FILE NAME Seq.zfill(2)_YYYYMMDD_NameWithoutSpaces
    file_extension = 'pdf' if '.pdf' in file_url else 'jpg'
    # add pdf convert?
    filename = f"{folder_location}/{str(index + 1).zfill(2)}_{date_stamp}_{title.title().replace(' ','')}.{file_extension}"
    request.urlretrieve(file_url, filename)

    # increment line count
    line += 1

# BLANK LINE
line += 1
# ADD LINE FOR SUM (c3/4)
sheet.cell(line, 3).value = 'Σ'
sheet.cell(line, 4).value = f"kr {round(total,2)}"
line += 1
# BLANK LINE
line+= 1
# LINE W/ NAME, F-NUMBER
sheet.cell(line, 1).value = 'Payable to:'
sheet.cell(line, 3).value = ''
sheet.cell(line, 4).value = ''
line += 1
# LINE W/ ACCOUNT NUMBER
sheet.cell(line, 3).value = '

# SAVE SPREADSHEET
wb.save('reimbursements.xlsx')
