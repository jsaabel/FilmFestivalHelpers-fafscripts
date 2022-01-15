# Rewrite 2.0 22.07.21
# added more properties,


from modules import notion as n, utils as u
import openpyxl as xl

# (convert ff export to .xlsx)

# load film freeway workbook
filmfreeway_file = '../files/FilmFreeway-Submissions.xlsx'
wb = xl.load_workbook(filmfreeway_file)
sheet = wb.active

# BUILD LIST AND DICTIONARY
properties = list()
credit_to_column = dict()
for number in range(1, sheet.max_column + 1):
    properties.append(sheet.cell(1, number).value)
    credit_to_column[sheet.cell(1, number).value] = number

# START FILM LOOP
for row in range(2, sheet.max_row + 1):
    # assign and clean film data
    film_data = dict()
    for property in properties:
        try:
            film_data[property] = sheet.cell(row, credit_to_column[property]).value
        except ValueError:
            film_data[property] = ""
    # DATA CLEANING
    if not film_data['Trailer URL']:  # avoid 400 response error for missing trailer url
        film_data['Trailer URL'] = None
    if film_data['Completion Date']:  # avoid 400 response error for missing year
        film_data['Completion Date'] = str(film_data['Completion Date'])[0:4]
    else:
        film_data['Completion Date'] = '(NO YEAR PROVIDED)'
    if film_data['If the film is selected, will it have a premiere during FAF 2021?'] == "Yes, Norwegian premiere":
        film_data['If the film is selected, will it have a premiere during FAF 2021?'] = "Norwegian premiere"
    elif film_data['If the film is selected, will it have a premiere during FAF 2021?'] == "Yes, World premiere":
        film_data['If the film is selected, will it have a premiere during FAF 2021?'] = "World premiere"
    else:
        film_data['If the film is selected, will it have a premiere during FAF 2021?'] = "No"
    if not film_data['Gender']:  # avoid 400 response error for missing gender
        film_data['Gender'] = 'Not provided'
    # convert comma separated cells to list objects
    film_data['Submission Categories'] = u.comma_separated_to_list(film_data['Submission Categories'])
    film_data['Country of Origin'] = u.comma_separated_to_list(film_data['Country of Origin'])
    film_data['Technique'] = u.comma_separated_to_list(film_data['Technique'])
    # calculate duration (min) as number
    duration = film_data['Duration'].split(':')
    minutes = float(duration[1])
    seconds = float(duration[2])
    film_data['Runtime_min'] = round(minutes + ((seconds + 10) / 60), 1)  # add a number of seconds for title etc
    # generate FilmFreeway link
    film_data['FilmFreeway Link'] = f"https://filmfreeway.com/submissions/{film_data['Submission ID']}"


    # WRITE TO FILM DB
    # NB: Properties must exist in db prior to import. See json for property names and types.
    # structure: [column name in spreadsheet][property name in notion][property type in notion]
    # building dictionary to send to notion API
    data_raw = dict()
    n.add_parent_id_to_request_dict(data_raw, parent_id=n.db_ids['test'])
    n.add_property_to_request_dict(data_raw, film_data['Project Title'], 'English Title', 'title')
    n.add_property_to_request_dict(data_raw, film_data['Project Title (Original Language)'], 'Original Title', 'text')
    n.add_property_to_request_dict(data_raw, film_data['Submission Categories'], 'Programme', 'multi_select')
    n.add_property_to_request_dict(data_raw, film_data['Directors'], 'Director', 'text')
    n.add_property_to_request_dict(data_raw, film_data['Submitter Biography'], 'Bio', 'text')
    n.add_property_to_request_dict(data_raw, film_data['If the film is selected, will it have a premiere during FAF 2021?'], 'Premiere', 'select')
    n.add_property_to_request_dict(data_raw, film_data['Completion Date'], 'Year', 'select')
    n.add_property_to_request_dict(data_raw, film_data['Gender'], 'Gender', 'select')
    n.add_property_to_request_dict(data_raw, film_data['Email'], 'Email', 'email')
    n.add_property_to_request_dict(data_raw, film_data['Country of Origin'], 'Country', 'multi_select')
    n.add_property_to_request_dict(data_raw, film_data['Synopsis'], 'Synopsis', 'text')
    n.add_property_to_request_dict(data_raw, film_data['Technique'], 'Technique', 'multi_select')
    n.add_property_to_request_dict(data_raw, film_data['Producers'], 'Production', 'text')
    n.add_property_to_request_dict(data_raw, film_data['Writers'], 'Screenwriter', 'text')
    n.add_property_to_request_dict(data_raw, film_data['Animator(s)'], 'Animation', 'text')
    n.add_property_to_request_dict(data_raw, film_data['Duration'][3:], 'Runtime', 'text')
    n.add_property_to_request_dict(data_raw, film_data['Runtime_min'], 'Runtime_min', 'number')
    n.add_property_to_request_dict(data_raw, film_data['Trailer URL'], 'Trailer URL', 'url')
    n.add_property_to_request_dict(data_raw, film_data['Submission Link'], 'Link', 'url')
    n.add_property_to_request_dict(data_raw, film_data['Submission Password'], 'Password', 'text')
    n.add_property_to_request_dict(data_raw, film_data['Submission ID'], 'Submission ID', 'text')
    n.add_property_to_request_dict(data_raw, film_data['FilmFreeway Link'], 'FilmFreeway Link', 'url')
    # sending to notion db
    print(f"Trying to write {film_data['Project Title']} to Notion...")
    n.write_to_db(data_dict=data_raw, db_id=n.db_ids['test'])



