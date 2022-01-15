# Image extraction from ff WIP
# last updated 22.07.2021 (probably not functional but ready to be adjusted for next year)
# ff links can/should be retrieved directly from notion db

from modules import notion as n
import re
from pynput.keyboard import Key, Controller
from time import sleep
import openpyxl as xl


def extract_img_links(filename: str):
    """returns a list of image links based found in provided html file (based on regex search string)"""
    with open(filename, 'r') as f:
        search_string = "[^;](https://storage.googleapis.com/ff-storage-p01\S*/\S*original/\S*\\.[jp][pn]g)"
        excludes = ['blurred', 'poster', 'headshot']
        img_links = list()
        results = re.findall(search_string, f.read())
        found_exceptions = list()
        for result in results:
            for exclude in excludes:
                if exclude in result:
                    found_exceptions.append(result)
        for found_exception in found_exceptions:
            results.remove(found_exception)
        for item in set(results):
            img_links.append(item)
        f.close()
    return img_links


# start up message
print('Script starting shortly. Open chrome tab now!')

# set up keyboard
kb = Controller()

# load workbook
wb = xl.load_workbook('FilmFreeway-Submissions.xlsx')
sheet = wb.active

# load notion db
notion_data = n.get_db('test')
item_count_notion = n.get_item_count(notion_data)
sleep(5)

# set up
"""here we gather all property names (column headings), store them in a list, and then store them 
together with their column number in a dictionary."""
properties = list()
credit_to_column = dict()
for number in range(1, sheet.max_column + 1):
    properties.append(sheet.cell(1, number).value)
    credit_to_column[sheet.cell(1, number).value] = number

# START LOOP

for row in range(2, sheet.max_row + 1):
    # assign film data
    film_data = dict()
    # building a dictionary for each entry from spreadsheet
    for property in properties:
        try:
            film_data[property] = sheet.cell(row, credit_to_column[property]).value
        except ValueError:
            film_data[property] = ""

    # extracting the relevant info from the previously created dictionary
    title = film_data['Project Title']
    id = film_data['Submission ID']
    url = film_data['ff_link']  # 'view-source:https://filmfreeway.com/submissions/18595893'  # (created in .xlsx)

    # SENDING KEYSTROKES
    sleep(1.5)
    # new tab
    with kb.pressed(Key.ctrl):
        kb.press('t')
        kb.release('t')
    sleep(1)
    # open url
    kb.type(url)
    sleep(0.2)
    kb.press(Key.enter)
    sleep(4)
    # save file
    with kb.pressed(Key.ctrl):
        kb.press('s')
        kb.release('s')
    sleep(1)
    kb.type(id)  # using the submission id as filename
    sleep(0.2)
    kb.press(Key.enter)
    sleep(1)
    # close tab
    with kb.pressed(Key.ctrl):
        kb.press('w')
        kb.release('w')

    # open file and extract urls
    filename = f"html test/{id}.html"
    try:
        links = extract_img_links(filename)
        # write url to spreadsheet
        sheet.cell(row, credit_to_column['extracted']).value = links[0]
    except (UnicodeDecodeError, IndexError):
        pass

    # match notion id and film
    for index in range(0, item_count_notion):
        if n.get_property(index, 'English Title', 'title', source=notion_data) == film_data['Project Title']:
            print(f"\t...match found: {film_data['Project Title']}")
            notion_id = n.get_property(index, property_type='id', source=notion_data)
            sheet.cell(row, credit_to_column['notion_id']).value = notion_id

            # create toggle block for urls in notion page
            n.write_block(parent_id=notion_id, content='Image links', block_type='toggle')
            # retrieve toggle block id
            toggle_id = n.get_children_ids(notion_id)[0][0]
            # write urls to notion
            for link in links:
                n.write_block(parent_id=toggle_id, content=link, block_type='link')

